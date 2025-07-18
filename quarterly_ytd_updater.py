#!/usr/bin/env python3
"""
Quarterly YTD Updater
Updates the quarterly YTD Excel file with completion data from all project lists.
"""

import pandas as pd
from datetime import datetime
import os
import re
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

def get_quarter_info():
    """
    Get quarter information from user input.
    Returns: dict with quarter details including file path, date range, months, etc.
    """
    print("="*60)
    print("QUARTERLY YTD UPDATER - QUARTER SELECTION")
    print("="*60)
    
    while True:
        print("\nSelect the quarter to process:")
        print("1. Q1 (January - March)")
        print("2. Q2 (April - June)")
        print("3. Q3 (July - September)")
        print("4. Q4 (October - December)")
        
        choice = input("\nEnter your choice (1-4): ").strip()
        
        if choice == '1':
            quarter_num = 1
            quarter_name = "1st Quarter"
            start_month, end_month = 1, 3
            month_names = ['January', 'February', 'March']
            month_abbrevs = ['Jan', 'Feb', 'Mar']
            break
        elif choice == '2':
            quarter_num = 2
            quarter_name = "2nd Quarter"
            start_month, end_month = 4, 6
            month_names = ['April', 'May', 'June']
            month_abbrevs = ['Apr', 'May', 'Jun']
            break
        elif choice == '3':
            quarter_num = 3
            quarter_name = "3rd Quarter"
            start_month, end_month = 7, 9
            month_names = ['July', 'August', 'September']
            month_abbrevs = ['Jul', 'Aug', 'Sep']
            break
        elif choice == '4':
            quarter_num = 4
            quarter_name = "4th Quarter"
            start_month, end_month = 10, 12
            month_names = ['October', 'November', 'December']
            month_abbrevs = ['Oct', 'Nov', 'Dec']
            break
        else:
            print("Invalid choice. Please enter 1, 2, 3, or 4.")
            continue
    
    # Get year
    while True:
        year_input = input(f"\nEnter the year for {quarter_name} (e.g., 2025): ").strip()
        try:
            year = int(year_input)
            if year < 2020 or year > 2030:
                print("Please enter a reasonable year between 2020 and 2030.")
                continue
            break
        except ValueError:
            print("Please enter a valid year.")
            continue
    
    # Calculate date range
    start_date = datetime(year, start_month, 1)
    if end_month == 12:
        end_date = datetime(year, 12, 31)
    elif end_month == 9:
        end_date = datetime(year, 9, 30)
    elif end_month == 6:
        end_date = datetime(year, 6, 30)
    else:  # Q1 - March
        end_date = datetime(year, 3, 31)
    
    # Setup quarterly sheets folder
    quarterly_sheets_dir = 'quarterly sheets'
    
    # Create the folder if it doesn't exist
    if not os.path.exists(quarterly_sheets_dir):
        os.makedirs(quarterly_sheets_dir, exist_ok=True)
        print(f"\n✓ Created 'quarterly sheets' folder")
    else:
        print(f"\n✓ Using existing 'quarterly sheets' folder")
    
    # Generate file path in quarterly sheets folder
    quarterly_file = os.path.join(quarterly_sheets_dir, f'{year} {quarter_name} YTD.xlsx')
    
    # Month indices for updating totals (0-based)
    month_indices = {}
    for i, month_num in enumerate(range(start_month, end_month + 1)):
        month_indices[month_num] = month_num - 1  # Convert to 0-based index
    
    quarter_info = {
        'quarter_num': quarter_num,
        'quarter_name': quarter_name,
        'year': year,
        'start_date': start_date,
        'end_date': end_date,
        'start_month': start_month,
        'end_month': end_month,
        'month_names': month_names,
        'month_abbrevs': month_abbrevs,
        'month_indices': month_indices,
        'quarterly_file': quarterly_file,
        'quarterly_sheets_dir': quarterly_sheets_dir,
        'project_lists': project_lists  # Use the global configuration
    }
    
    print(f"\n✓ Selected: {quarter_name} {year}")
    print(f"✓ Date range: {start_date.strftime('%B %d, %Y')} - {end_date.strftime('%B %d, %Y')}")
    print(f"✓ Target file: {quarterly_file}")
    
    return quarter_info


## Project list files
project_lists = [
    ('2023', r'N:\Project List\2023 Project List\2023 Project List.xlsx'),
    ('2024', r'N:\Project List\2024 Project List\2024 Project List.xlsx'),
    ('2025', r'N:\Project List\2025 Project List\2025 Project List.xlsx'),
]

# Add logic to use backup from 'quarterly sheets' if N: drive file is not found
for idx, (year, n_drive_path) in enumerate(project_lists):
    if not os.path.exists(n_drive_path):
        backup_path = os.path.join('quarterly sheets', os.path.basename(n_drive_path))
        if os.path.exists(backup_path):
            print(f"N: drive file for {year} not found, using backup in quarterly sheets: {backup_path}")
            project_lists[idx] = (year, backup_path)
        else:
            print(f"Warning: Neither N: drive nor backup found for {year} Project List: {n_drive_path}")

def parse_split_invoices(comments_text, original_amount):
    """
    Parse split invoice information from comments text.
    Returns a list of tuples: (percentage, date, description)
    """
    if pd.isna(comments_text) or not str(comments_text).strip():
        return []
    
    comments_str = str(comments_text).strip()
    
    # Pattern to match format: "Invoiced at 50% on 4/18/2025. Invoiced rest on 5/13/2025"
    # First, look for the initial percentage and date
    initial_pattern = r'Invoiced at (\d+(?:\.\d+)?)%\s*on\s*(\d{1,2}[/-]\d{1,2}[/-]\d{4})'
    initial_match = re.search(initial_pattern, comments_str, re.IGNORECASE)
    
    if not initial_match:
        # Fallback to old pattern for backwards compatibility
        old_pattern = r'(\d+(?:\.\d+)?)%\s*(?:invoiced|billed|paid)\s*(\d{1,2}[/-]\d{1,2}[/-]\d{4})'
        matches = re.findall(old_pattern, comments_str, re.IGNORECASE)
        
        if not matches:
            return []
        
        split_invoices = []
        for percentage_str, date_str in matches:
            try:
                percentage = float(percentage_str)
                # Parse date with flexible format
                if '/' in date_str:
                    date_parts = date_str.split('/')
                else:
                    date_parts = date_str.split('-')
                
                if len(date_parts) == 3:
                    month, day, year = date_parts
                    parsed_date = datetime(int(year), int(month), int(day))
                    
                    description = f"{percentage}% invoiced {parsed_date.strftime('%m/%d/%Y')}"
                    split_invoices.append((percentage, parsed_date, description))
            except ValueError as e:
                print(f"Warning: Could not parse split invoice '{percentage_str}% {date_str}' from comments: {e}")
                continue
        
        # Validate percentages add up reasonably (allow some tolerance for rounding)
        total_percentage = sum(p[0] for p in split_invoices)
        if abs(total_percentage - 100.0) > 1.0:  # Allow 1% tolerance for rounding
            print(f"Warning: Split invoice percentages don't add up to 100%: {total_percentage}% in '{comments_str}'")
        
        return split_invoices
    
    # New format processing
    split_invoices = []
    
    try:
        # Parse the initial percentage and date
        initial_percentage = float(initial_match.group(1))
        initial_date_str = initial_match.group(2)
        
        # Parse the initial date
        if '/' in initial_date_str:
            date_parts = initial_date_str.split('/')
        else:
            date_parts = initial_date_str.split('-')
        
        if len(date_parts) == 3:
            month, day, year = date_parts
            initial_date = datetime(int(year), int(month), int(day))
            
            initial_description = f"{initial_percentage}% invoiced {initial_date.strftime('%m/%d/%Y')}"
            split_invoices.append((initial_percentage, initial_date, initial_description))
        
        # Look for the "rest" date
        rest_pattern = r'Invoiced rest on\s*(\d{1,2}[/-]\d{1,2}[/-]\d{4})'
        rest_match = re.search(rest_pattern, comments_str, re.IGNORECASE)
        
        if rest_match:
            rest_date_str = rest_match.group(1)
            
            # Parse the rest date
            if '/' in rest_date_str:
                date_parts = rest_date_str.split('/')
            else:
                date_parts = rest_date_str.split('-')
            
            if len(date_parts) == 3:
                month, day, year = date_parts
                rest_date = datetime(int(year), int(month), int(day))
                
                # Calculate the remaining percentage
                rest_percentage = 100.0 - initial_percentage
                
                rest_description = f"{rest_percentage}% invoiced {rest_date.strftime('%m/%d/%Y')} (rest)"
                split_invoices.append((rest_percentage, rest_date, rest_description))
        
    except ValueError as e:
        print(f"Warning: Could not parse new format split invoice from comments '{comments_str}': {e}")
        return []
    
    return split_invoices

def create_split_records(original_record, split_invoices, original_amount, amount_invoiced):
    """
    Create multiple records from a single record with split invoicing.
    For new format: first record gets amount_invoiced, rest gets (original_amount - amount_invoiced)
    Total across all records should equal original_amount (the "Amount" column)
    Returns a list of modified records.
    """
    if not split_invoices:
        return [original_record]
    
    split_records = []
    
    for i, (percentage, invoice_date, description) in enumerate(split_invoices):
        # Create a copy of the original record
        split_record = original_record.copy()
        
        # For new format, calculate amount to ensure total equals original_amount
        if len(split_invoices) == 2 and "rest" in description.lower():
            # This is the "rest" invoice - use remaining amount to reach original_amount total
            proportional_amount = original_amount - amount_invoiced
        elif len(split_invoices) == 2 and i == 0:
            # This is the first invoice in new format - use amount_invoiced
            proportional_amount = amount_invoiced
        else:
            # Fallback to percentage calculation based on original_amount (old format or other cases)
            proportional_amount = original_amount * (percentage / 100.0)
        
        # Update the record with split invoice information
        split_record['Invoice Date'] = invoice_date
        split_record['Amount Invoiced'] = proportional_amount
        split_record['Split Invoice Description'] = description
        split_record['Original Amount'] = original_amount
        split_record['Split Percentage'] = percentage
        
        split_records.append(split_record)
    
    # Verify total equals original_amount
    total_split = sum(record['Amount Invoiced'] for record in split_records)
    if abs(total_split - original_amount) > 0.01:  # Small tolerance for floating point
        print(f"Warning: Split invoice total (${total_split:,.2f}) doesn't equal original amount (${original_amount:,.2f})")
    
    return split_records

def collect_completion_data(quarter_info):
    """
    Collect all completion data from project lists across all years.
    Returns a DataFrame with completion dates and amounts.
    """
    print(f"Collecting completion data from project lists for {quarter_info['quarter_name']} {quarter_info['year']}...")
    
    all_completion_data = []
    
    for year, file_path in quarter_info['project_lists']:
        print(f"\nProcessing {year} Project List...")
        
        try:
            # Read the project list
            df = pd.read_excel(file_path, sheet_name=year, header=5)
            
            # Handle different ACGI column names across years
            acgi_col = None
            for col in df.columns:
                if 'acgi' in str(col).lower() and '#' in str(col):
                    acgi_col = col
                    break
            
            if acgi_col is None:
                print(f"Warning: Could not find ACGI column in {year}, using first column")
                acgi_col = df.columns[0]
            
            # Handle different project name column variations
            project_col = None
            for col in df.columns:
                if 'project' in str(col).lower() and ('number' in str(col).lower() or 'name' in str(col).lower()):
                    project_col = col
                    break
            if project_col is None:
                project_col = 'Project Number/Name'  # fallback
            
            # Handle different client PO column variations
            client_col = None
            for col in df.columns:
                if 'client' in str(col).lower() and 'po' in str(col).lower():
                    client_col = col
                    break
            if client_col is None:
                client_col = 'Client / PO #'  # fallback
            
            # Handle different Line # column variations
            line_col = None
            for col in df.columns:
                if 'line' in str(col).lower() and '#' in str(col):
                    line_col = col
                    break
            if line_col is None:
                line_col = 'Line #'  # fallback
            
            # Handle different PO Date column variations  
            po_date_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'po' in col_lower and 'date' in col_lower:
                    po_date_col = col
                    break
            if po_date_col is None:
                po_date_col = 'PO Date'  # fallback
            
            # Handle different Dept column variations
            dept_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'dept' in col_lower or 'department' in col_lower:
                    dept_col = col
                    break
            if dept_col is None:
                dept_col = 'Dept'  # fallback
            
            # Handle different Invoice Date column variations
            invoice_date_col = None
            for col in df.columns:
                col_lower = str(col).lower()
                if 'invoice' in col_lower and 'date' in col_lower:
                    invoice_date_col = col
                    break
            if invoice_date_col is None:
                invoice_date_col = 'Invoice Date'  # fallback
            
            # Handle different Comments column variations
            comments_col = None
            
            # First try to find column N specifically (Comments field is in column N)
            if len(df.columns) >= 14:  # Column N is the 14th column (0-indexed = 13)
                comments_col = df.columns[13]  # Column N (0-indexed)
                print(f"Using column N ({comments_col}) as Comments field")
            else:
                # Fallback: search by name if column N doesn't exist
                for col in df.columns:
                    col_lower = str(col).lower()
                    if 'comment' in col_lower or 'note' in col_lower:
                        comments_col = col
                        break
                if comments_col is None:
                    comments_col = 'Comments'  # fallback
                print(f"Column N not available, using {comments_col} as Comments field")
            
            # Select columns with fallback handling
            required_cols = ['Amount Invoiced', 'Amount']  # Need both Amount and Amount Invoiced for split calculation
            optional_cols = [acgi_col, project_col, client_col, line_col, po_date_col, dept_col, invoice_date_col, 'Completion Date', comments_col]
            
            # Check which columns actually exist
            available_cols = required_cols.copy()
            for col in optional_cols:
                if col in df.columns:
                    available_cols.append(col)
                else:
                    print(f"Warning: Column '{col}' not found in {year}")
            
            # Debug: Print all available columns to help identify columns
            print(f"Available columns in {year}: {list(df.columns)}")
            
            completion_data = df[available_cols].copy()
            
            # Rename columns to standardize
            rename_dict = {acgi_col: 'ACGI #'}
            if project_col in df.columns:
                rename_dict[project_col] = 'Project Number/Name'
            if client_col in df.columns:
                rename_dict[client_col] = 'Client / PO #'
            if line_col in df.columns:
                rename_dict[line_col] = 'Line #'
            if po_date_col in df.columns:
                rename_dict[po_date_col] = 'PO Date'
            if dept_col in df.columns:
                rename_dict[dept_col] = 'Dept'
            if invoice_date_col in df.columns:
                rename_dict[invoice_date_col] = 'Invoice Date'
            if comments_col in df.columns:
                rename_dict[comments_col] = 'Comments'
            
            completion_data = completion_data.rename(columns=rename_dict)
            
            # Clean the data - require both Amount and Amount Invoiced for split invoice calculations
            completion_data = completion_data.dropna(subset=['Amount Invoiced', 'Amount'])
            
            # Debug: Check for specific record before any filtering
            acgi_col_name = acgi_col if acgi_col in df.columns else 'ACGI #'
            if acgi_col_name in df.columns:
                try:
                    # Check for 25-1376
                    mask = df[acgi_col_name].astype(str).str.contains('25-1376', na=False)
                    specific_record = df[mask]
                    if not specific_record.empty:
                        record = specific_record.iloc[0]
                        print(f"  *** Found 25-1376 in raw data: ***")
                        print(f"      ACGI #: {record.get(acgi_col_name, 'N/A')}")
                        print(f"      Invoice Date (raw): {record.get(invoice_date_col, 'N/A')}")
                        print(f"      Completion Date (raw): {record.get('Completion Date', 'N/A')}")
                        print(f"      Amount Invoiced (raw): {record.get('Amount Invoiced', 'N/A')}")
                        print(f"      Comments (raw): {record.get(comments_col, 'N/A')}")
                    
                    # Check for 24-3163
                    mask_3163 = df[acgi_col_name].astype(str).str.contains('24-3163', na=False)
                    specific_record_3163 = df[mask_3163]
                    if not specific_record_3163.empty:
                        record = specific_record_3163.iloc[0]
                        print(f"  *** Found 24-3163 in raw data: ***")
                        print(f"      ACGI #: {record.get(acgi_col_name, 'N/A')}")
                        print(f"      Invoice Date (raw): {record.get(invoice_date_col, 'N/A')}")
                        print(f"      Completion Date (raw): {record.get('Completion Date', 'N/A')}")
                        print(f"      Amount Invoiced (raw): {record.get('Amount Invoiced', 'N/A')}")
                        print(f"      Comments (raw): {record.get(comments_col, 'N/A')}")
                    else:
                        print(f"  *** 24-3163 NOT found in raw {year} data ***")
                except Exception as e:
                    print(f"  Debug search for records failed: {e}")
            
            # Handle problematic date entries
            def safe_date_parse(date_val):
                if pd.isna(date_val):
                    return pd.NaT
                if isinstance(date_val, str) and date_val.strip() == '':
                    return pd.NaT
                try:
                    parsed_date = pd.to_datetime(date_val)
                    return parsed_date
                except:
                    print(f"Warning: Could not parse date '{date_val}' in {year}")
                    return pd.NaT
            
            # Parse Completion Date (optional - we don't filter on this anymore)
            print(f"Parsing completion dates in {year} (optional)...")
            completion_data['Completion Date'] = completion_data['Completion Date'].apply(safe_date_parse)
            
            # Parse Invoice Date for filtering (REQUIRED)
            print(f"Parsing invoice dates in {year} (required)...")
            completion_data['Invoice Date'] = completion_data['Invoice Date'].apply(safe_date_parse)
            
            # Clean and prepare data for split invoice processing
            completion_data['Amount Invoiced'] = pd.to_numeric(completion_data['Amount Invoiced'], errors='coerce')
            print(f"Records before amount filtering: {len(completion_data)}")
            completion_data = completion_data.dropna(subset=['Amount Invoiced'])  # Remove non-numeric amounts
            print(f"Records after amount filtering: {len(completion_data)}")
            
            # *** SPLIT INVOICE PROCESSING ***
            print(f"Processing split invoices from comments in {year}...")
            expanded_records = []
            split_count = 0
            
            for idx, row in completion_data.iterrows():
                original_amount = row['Amount']  # Total project amount - this is what should go in the sheet
                amount_invoiced = row['Amount Invoiced']  # Amount actually invoiced so far
                comments = row.get('Comments', '')
                
                # Safely handle comments that might be NaN or other non-string types
                if pd.isna(comments) or not isinstance(comments, str):
                    comments = ''
                else:
                    comments = str(comments).strip()
                
                # First check if there are split invoice patterns in comments
                split_invoices = parse_split_invoices(comments, original_amount)
                
                if split_invoices:
                    # Split comments found - now check if we need to split based on Amount vs Amount Invoiced
                    should_split = abs(original_amount - amount_invoiced) > 0.01  # Use small tolerance for floating point comparison
                    
                    if should_split:
                        # Create multiple records for split invoices
                        split_records = create_split_records(row, split_invoices, original_amount, amount_invoiced)
                        expanded_records.extend(split_records)
                        split_count += 1
                        
                        # Debug: Show split invoice processing
                        acgi_num = row.get('ACGI #', 'Unknown')
                        print(f"  Split invoice found for {acgi_num}:")
                        print(f"    Total project amount: ${original_amount:,.2f}")
                        print(f"    Amount invoiced so far: ${amount_invoiced:,.2f}")
                        print(f"    Difference: ${original_amount - amount_invoiced:,.2f}")
                        print(f"    Comments: {comments}")
                        for i, (percentage, date, desc) in enumerate(split_invoices):
                            if len(split_invoices) == 2 and "rest" in desc.lower():
                                amount = original_amount - amount_invoiced
                            elif len(split_invoices) == 2 and i == 0:
                                amount = amount_invoiced
                            else:
                                amount = original_amount * (percentage / 100.0)
                            print(f"    Split {i+1}: {desc} = ${amount:,.2f} on {date.strftime('%m/%d/%Y')}")
                    else:
                        # Split comments but Amount = Amount Invoiced, use full Amount on Invoice Date
                        row_copy = row.copy()
                        row_copy['Amount Invoiced'] = original_amount  # Use full Amount value
                        expanded_records.append(row_copy)
                        print(f"  {row.get('ACGI #', 'Unknown')}: Split comments but Amount Invoiced = Amount (${original_amount:,.2f}), using full amount on Invoice Date")
                else:
                    # No split comments found - just use Amount Invoiced as-is, regardless of Amount mismatch
                    expanded_records.append(row)
                    
                    # Optional debug for large mismatches (but we won't change the amount)
                    if abs(original_amount - amount_invoiced) > 0.01:
                        acgi_num = row.get('ACGI #', 'Unknown')
                        print(f"  {acgi_num}: No split comments, using Amount Invoiced (${amount_invoiced:,.2f}) despite Amount being (${original_amount:,.2f})")
            
            # Convert back to DataFrame
            completion_data = pd.DataFrame(expanded_records)
            print(f"Split invoice processing complete: {split_count} records split into {len(completion_data)} total records")
            
            # Now filter by Invoice Date after split processing
            print(f"Records before invoice date filtering: {len(completion_data)}")
            completion_data = completion_data.dropna(subset=['Invoice Date'])  # Remove unparseable invoice dates
            print(f"Records after invoice date filtering: {len(completion_data)}")
            
            # Filter for the selected quarter dates based on INVOICE DATE
            quarter_start = quarter_info['start_date']
            quarter_end = quarter_info['end_date']
            quarter_name = quarter_info['quarter_name']
            year_selected = quarter_info['year']
            
            print(f"Records before {quarter_name} date filtering: {len(completion_data)}")
            
            # Debug: Check for specific records before quarter filtering
            specific_record_before_quarter = completion_data[completion_data.get('ACGI #', pd.Series()).astype(str).str.contains('25-1376', na=False)]
            if not specific_record_before_quarter.empty:
                record = specific_record_before_quarter.iloc[0]
                invoice_date = record.get('Invoice Date', 'N/A')
                print(f"  *** 25-1376 before {quarter_name} filtering: ***")
                print(f"      Invoice Date: {invoice_date}")
                print(f"      Is in {quarter_name} range ({quarter_start.date()} to {quarter_end.date()}): {quarter_start <= invoice_date <= quarter_end if pd.notna(invoice_date) else 'N/A'}")
            
            # Debug: Check for 24-3163 before quarter filtering
            specific_record_3163_before_quarter = completion_data[completion_data.get('ACGI #', pd.Series()).astype(str).str.contains('24-3163', na=False)]
            if not specific_record_3163_before_quarter.empty:
                record = specific_record_3163_before_quarter.iloc[0]
                invoice_date = record.get('Invoice Date', 'N/A')
                print(f"  *** 24-3163 before {quarter_name} filtering: ***")
                print(f"      Invoice Date: {invoice_date}")
                print(f"      Is in {quarter_name} range ({quarter_start.date()} to {quarter_end.date()}): {quarter_start <= invoice_date <= quarter_end if pd.notna(invoice_date) else 'N/A'}")
            else:
                print(f"  *** 24-3163 not found before {quarter_name} filtering ***")
            
            quarter_data = completion_data[
                (completion_data['Invoice Date'] >= quarter_start) & 
                (completion_data['Invoice Date'] <= quarter_end)
            ].copy()
            
            print(f"Records after {quarter_name} date filtering: {len(quarter_data)}")
            
            # Debug: Check for 25-1376 after quarter filtering
            specific_record_after_quarter = quarter_data[quarter_data.get('ACGI #', pd.Series()).astype(str).str.contains('25-1376', na=False)]
            if not specific_record_after_quarter.empty:
                print(f"  *** 25-1376 found in final {quarter_name} data ***")
            elif not specific_record_before_quarter.empty:
                print(f"  *** 25-1376 was filtered OUT by {quarter_name} date range ***")
            
            if not quarter_data.empty:
                quarter_data['Source_Year'] = year
                all_completion_data.append(quarter_data)
                print(f"Found {len(quarter_data)} {quarter_name} {year_selected} completion records from {year}")
            else:
                print(f"No {quarter_name} {year_selected} completion records found in {year}")
                
        except Exception as e:
            print(f"Error processing {year} Project List: {e}")
    
    if all_completion_data:
        combined_data = pd.concat(all_completion_data, ignore_index=True)
        print(f"\nTotal {quarter_info['quarter_name']} {quarter_info['year']} completion records collected: {len(combined_data)}")
        return combined_data
    else:
        print(f"\nNo completion data found for {quarter_info['quarter_name']} {quarter_info['year']}")
        return pd.DataFrame()

def calculate_monthly_totals(completion_data):
    """
    Calculate monthly totals from completion data.
    """
    if completion_data.empty:
        return {}
    
    # Group by month and sum amounts
    completion_data['Month'] = completion_data['Completion Date'].dt.month
    monthly_totals = completion_data.groupby('Month')['Amount Invoiced'].sum()
    
    # Convert to dictionary with month names
    month_names = {4: 'April', 5: 'May', 6: 'June'}
    totals_dict = {}
    
    for month_num, total in monthly_totals.items():
        if month_num in month_names:
            totals_dict[month_names[month_num]] = total
    
    return totals_dict

def update_quarterly_ytd(completion_data, quarter_info):
    """
    Update the quarterly YTD Excel file with new completion data and detailed formatting.
    """
    quarterly_file = quarter_info['quarterly_file']
    print(f"\nUpdating {quarterly_file}...")
    
    # Read existing monthly totals from previous quarter file or current file if it exists
    existing_monthly_totals = [0] * 13  # Initialize all to 0
    
    # Determine previous quarter file to read totals from
    previous_quarter_file = None
    if quarter_info['quarter_num'] > 1:
        # Map quarter numbers to names
        quarter_names = {1: "1st Quarter", 2: "2nd Quarter", 3: "3rd Quarter", 4: "4th Quarter"}
        previous_quarter_name = quarter_names[quarter_info['quarter_num'] - 1]
        previous_quarter_file = os.path.join(quarter_info['quarterly_sheets_dir'], 
                                           f"{quarter_info['year']} {previous_quarter_name} YTD.xlsx")
    
    # Try to read from previous quarter file first
    if previous_quarter_file and os.path.exists(previous_quarter_file):
        print(f"Reading monthly totals from previous quarter: {previous_quarter_file}")
        try:
            previous_wb = load_workbook(previous_quarter_file)
            previous_ws = previous_wb.active
            
            # Read monthly totals from row 2 (columns 1-13)
            for col in range(1, 14):  # Columns A-M (1-13)
                try:
                    cell_value = previous_ws.cell(row=2, column=col).value
                    if cell_value is not None and isinstance(cell_value, (int, float)):
                        existing_monthly_totals[col-1] = float(cell_value)
                        if col <= 12:  # Don't show YTD in the list
                            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                                         'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
                            print(f"  {month_names[col-1]}: ${cell_value:,.2f}")
                except Exception as e:
                    print(f"  Warning: Could not read previous quarter value for column {col}: {e}")
            
            previous_wb.close()
            print(f"✓ Successfully loaded totals from previous quarter")
            
        except Exception as e:
            print(f"Warning: Could not read previous quarter file: {e}")
            print("Will start with zero values for all months")
    elif quarter_info['quarter_num'] > 1:
        print(f"Previous quarter file not found: {previous_quarter_file}")
        print("Using hardcoded Q1 values as baseline:")
        # Hardcode Q1 2025 values when previous quarter file is not available
        existing_monthly_totals[0] = 872459.74   # January
        existing_monthly_totals[1] = 609301.81   # February  
        existing_monthly_totals[2] = 463345.08   # March
        month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                      'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
        for i in range(3):  # Show Q1 months
            print(f"  {month_names[i]}: ${existing_monthly_totals[i]:,.2f}")
    else:
        print("Q1 - no previous quarter, starting with zero values")
    
    # Create backup of current quarter file if it exists
    if os.path.exists(quarterly_file):
        backup_file = quarterly_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        try:
            import shutil
            shutil.copy2(quarterly_file, backup_file)
            print(f"Backup of current file saved as: {backup_file}")
        except Exception as backup_error:
            print(f"Warning: Could not create backup: {backup_error}")
    
    # Create new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = f"Q{quarter_info['quarter_num']} {quarter_info['year']} YTD"
    
    # Define styles
    # Month header style (bold, green background)
    month_style = Font(bold=True, color="FFFFFF")
    month_fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
    
    # Total style (bold, red font)
    total_style = Font(bold=True, color="FF0000")
    
    # Day header style (bold, black border)
    day_style = Font(bold=True)
    day_border = Border(
        left=Side(border_style="thick", color="000000"),
        right=Side(border_style="thick", color="000000"),
        top=Side(border_style="thick", color="000000"),
        bottom=Side(border_style="thick", color="000000")
    )
    
    # Column header style (bold, gray background)
    header_style = Font(bold=True)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    # Regular border
    regular_border = Border(
        left=Side(border_style="thin", color="000000"),
        right=Side(border_style="thin", color="000000"),
        top=Side(border_style="thin", color="000000"),
        bottom=Side(border_style="thin", color="000000")
    )
    
    # Center alignment
    center_align = Alignment(horizontal='center', vertical='center')
    
    # Current row tracker
    current_row = 1
    
    # --- HEADER SECTION ---
    # Row 1: Month headers
    month_names = ['January', 'February', 'March', 'April', 'May', 'June', 
                  'July', 'August', 'September', 'October', 'November', 'December', 'YTD Totals']
    
    for col, month in enumerate(month_names, 1):
        cell = ws.cell(row=current_row, column=col, value=month)
        cell.font = month_style
        cell.fill = month_fill
        cell.alignment = center_align
        cell.border = regular_border
    
    current_row += 1
    
    # Row 2: Monthly totals (calculate Q2 totals, preserve existing values for other months)
    daily_totals = {}
    if not completion_data.empty:
        # Group by date and calculate daily totals
        completion_data_sorted = completion_data.sort_values('Invoice Date').reset_index(drop=True)
        completion_data_sorted['Date_Only'] = completion_data_sorted['Invoice Date'].dt.date
        grouped_by_date = completion_data_sorted.groupby('Date_Only')
        
        # Calculate monthly totals from daily totals
        for date_only, date_group in grouped_by_date:
            month = date_only.month
            daily_total = date_group['Amount Invoiced'].sum()
            if month not in daily_totals:
                daily_totals[month] = 0
            daily_totals[month] += daily_total
    
    # Start with existing monthly totals and only update the selected quarter months
    monthly_totals = existing_monthly_totals.copy()
    
    # Update only the selected quarter months with calculated values
    print(f"Updating {quarter_info['quarter_name']} monthly totals:")
    month_indices = quarter_info['month_indices']
    for month_num, total in daily_totals.items():
        if month_num in month_indices:
            old_value = monthly_totals[month_indices[month_num]]
            monthly_totals[month_indices[month_num]] = total
            month_names = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 
                          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
            print(f"  {month_names[month_indices[month_num]]}: ${old_value:,.2f} → ${total:,.2f}")
    
    # Calculate YTD total from all months (existing + updated)
    ytd_total = sum(monthly_totals[:12])  # Sum first 12 months, exclude old YTD
    monthly_totals[12] = ytd_total
    print(f"  YTD Total: ${ytd_total:,.2f}")
    
    # Write monthly totals with formatting
    for col, total in enumerate(monthly_totals, 1):
        cell = ws.cell(row=current_row, column=col, value=total)
        if total > 0:
            cell.number_format = '"$"#,##0.00'
            cell.font = total_style
        cell.alignment = center_align
        cell.border = regular_border
    
    current_row += 1
    
    # Empty separator rows
    for _ in range(3):
        current_row += 1
    
    # --- DATA SECTION ---
    if not completion_data.empty:
        print(f"\nAdding {len(completion_data)} completion records with formatting...")
        
        # Sort and group completion data by invoice date
        completion_data_sorted = completion_data.sort_values('Invoice Date').reset_index(drop=True)
        completion_data_sorted['Date_Only'] = completion_data_sorted['Invoice Date'].dt.date
        grouped_by_date = completion_data_sorted.groupby('Date_Only')
        
        date_groups = list(grouped_by_date)
        
        for i, (date_only, date_group) in enumerate(date_groups):
            # Add date header
            day_of_week = date_only.strftime('%A')
            date_header = f"{day_of_week} {date_only.month}-{date_only.day}-{date_only.year} (Invoice Date)"
            
            date_cell = ws.cell(row=current_row, column=1, value=date_header)
            date_cell.font = day_style
            date_cell.border = day_border
            date_cell.alignment = center_align
            
            # Merge cells for date header (columns A-J)
            ws.merge_cells(f'A{current_row}:J{current_row}')
            current_row += 1
            
            # Add column headers for this day
            headers = ['ACGI Project/ Invoice #', 'Dept', 'Project Number/ Name', 'Type', 
                      'Client / PO #', 'Line #', 'PO Date', 'Amount', 'Invoice Date', 'Amount Invoiced']
            
            for col, header in enumerate(headers, 1):
                header_cell = ws.cell(row=current_row, column=col, value=header)
                header_cell.font = header_style
                header_cell.fill = header_fill
                header_cell.alignment = center_align
                header_cell.border = regular_border
            
            current_row += 1
            
            # Sort this day's transactions by Project Number
            date_group_sorted = date_group.copy()
            
            def parse_project_number_for_sort(project_num):
                if pd.isna(project_num) or project_num == '':
                    return (0, 0)
                try:
                    project_str = str(project_num).strip()
                    if '-' in project_str:
                        year_part, num_part = project_str.split('-', 1)
                        year = int(year_part)
                        num = int(num_part.lstrip('0') or '0')
                        return (year, num)
                    else:
                        return (0, int(project_str))
                except:
                    return (0, 0)
            
            date_group_sorted['Project_Sort_Key'] = date_group_sorted['ACGI #'].apply(parse_project_number_for_sort)
            date_group_sorted = date_group_sorted.sort_values('Project_Sort_Key')
            
            # Add transaction rows
            daily_total = 0
            for _, row in date_group_sorted.iterrows():
                # Column 1: ACGI #
                ws.cell(row=current_row, column=1, value=str(row.get('ACGI #', ''))).border = regular_border
                
                # Column 2: Dept
                ws.cell(row=current_row, column=2, value=str(row.get('Dept', ''))).border = regular_border
                
                # Column 3: Project Number/Name (with split info if applicable)
                project_name = str(row.get('Project Number/Name', ''))
                if 'Split Invoice Description' in row and pd.notna(row['Split Invoice Description']):
                    project_name += f" [{row['Split Invoice Description']}]"
                ws.cell(row=current_row, column=3, value=project_name).border = regular_border
                
                # Column 4: Type
                ws.cell(row=current_row, column=4, value=str(row.get('Type', 'Completion'))).border = regular_border
                
                # Column 5: Client/PO #
                ws.cell(row=current_row, column=5, value=str(row.get('Client / PO #', ''))).border = regular_border
                
                # Column 6: Line #
                ws.cell(row=current_row, column=6, value=str(row.get('Line #', ''))).border = regular_border
                
                # Column 7: PO Date
                po_date = row.get('PO Date', '')
                if pd.notna(po_date) and po_date != '':
                    try:
                        if isinstance(po_date, str):
                            po_date_parsed = pd.to_datetime(po_date)
                        else:
                            po_date_parsed = po_date
                        po_date_str = po_date_parsed.strftime('%m/%d/%Y')
                    except:
                        po_date_str = str(po_date)
                else:
                    po_date_str = ''
                ws.cell(row=current_row, column=7, value=po_date_str).border = regular_border
                
                # Column 8: Amount (with currency formatting)
                amount = float(row['Amount Invoiced']) if pd.notna(row['Amount Invoiced']) else 0
                amount_cell = ws.cell(row=current_row, column=8, value=amount)
                amount_cell.number_format = '"$"#,##0.00'
                amount_cell.border = regular_border
                daily_total += amount
                
                # Column 9: Invoice Date
                try:
                    invoice_date = row['Invoice Date']
                    invoice_date_str = invoice_date.strftime('%m/%d/%Y')
                except:
                    invoice_date_str = str(row['Invoice Date'])
                ws.cell(row=current_row, column=9, value=invoice_date_str).border = regular_border
                
                # Column 10: Amount Invoiced (with currency formatting)
                amount_invoiced_cell = ws.cell(row=current_row, column=10, value=amount)
                amount_invoiced_cell.number_format = '"$"#,##0.00'
                amount_invoiced_cell.border = regular_border
                
                current_row += 1
            
            # Add daily total row
            total_label_cell = ws.cell(row=current_row, column=1, value=f"Daily Total for {date_only.strftime('%m/%d/%Y')} (Invoice Date)")
            total_label_cell.font = total_style
            total_label_cell.border = regular_border
            
            # Amount total
            amount_total_cell = ws.cell(row=current_row, column=8, value=daily_total)
            amount_total_cell.number_format = '"$"#,##0.00'
            amount_total_cell.font = total_style
            amount_total_cell.border = regular_border
            
            # Amount Invoiced total
            amount_invoiced_total_cell = ws.cell(row=current_row, column=10, value=daily_total)
            amount_invoiced_total_cell.number_format = '"$"#,##0.00'
            amount_invoiced_total_cell.font = total_style
            amount_invoiced_total_cell.border = regular_border
            
            # Empty cells for other columns
            for col in [2, 3, 4, 5, 6, 7, 9]:
                ws.cell(row=current_row, column=col).border = regular_border
            
            current_row += 1
            
            # Add empty row between days (except for last day)
            if i < len(date_groups) - 1:
                current_row += 1
    
    # Auto-adjust column widths
    for col in range(1, 14):  # Extended to include all 13 columns (YTD column is 13)
        column_letter = get_column_letter(col)
        max_length = 0
        
        for row in ws[column_letter]:
            try:
                if len(str(row.value)) > max_length:
                    max_length = len(str(row.value))
            except:
                pass
        
        # Set minimum width and add padding, with extra width for YTD column
        if col == 13:  # YTD column needs extra width for large numbers
            adjusted_width = max(max_length + 2, 18)  # Minimum 18 for YTD
        else:
            adjusted_width = max(max_length + 2, 12)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save the workbook
    try:
        print("Saving formatted file...")
        wb.save(quarterly_file)
        print(f"✓ Successfully updated {quarterly_file} with formatting")
        return True
        
    except Exception as e:
        print(f"Error saving file: {e}")
        
        # Try saving with a different name
        try:
            alt_file = quarterly_file.replace('.xlsx', '_formatted.xlsx')
            wb.save(alt_file)
            print(f"✓ Saved as {alt_file}")
            return True
        except Exception as final_error:
            print(f"Final attempt failed: {final_error}")
            return False

def print_summary(completion_data):
    """
    Print a summary of the completion data.
    """
    if completion_data.empty:
        print("\nNo completion data to summarize.")
        return
    
    print("\n" + "="*60)
    print("COMPLETION DATA SUMMARY")
    print("="*60)
    
    # Summary by month
    monthly_summary = completion_data.groupby(completion_data['Invoice Date'].dt.strftime('%B %Y')).agg({
        'Amount Invoiced': ['count', 'sum']
    }).round(2)
    
    print("\nBy Month:")
    for month in monthly_summary.index:
        count = monthly_summary.loc[month, ('Amount Invoiced', 'count')]
        total = monthly_summary.loc[month, ('Amount Invoiced', 'sum')]
        print(f"  {month}: {count} records, ${total:,.2f}")
    
    # Summary by source year
    print("\nBy Source Year:")
    source_summary = completion_data.groupby('Source_Year').agg({
        'Amount Invoiced': ['count', 'sum']
    }).round(2)
    
    for year in source_summary.index:
        count = source_summary.loc[year, ('Amount Invoiced', 'count')]
        total = source_summary.loc[year, ('Amount Invoiced', 'sum')]
        print(f"  {year}: {count} records, ${total:,.2f}")
    
    # Overall total
    total_amount = completion_data['Amount Invoiced'].sum()
    total_count = len(completion_data)
    print(f"\nOverall Total: {total_count} records, ${total_amount:,.2f}")
    
    # Split invoice summary
    if 'Split Invoice Description' in completion_data.columns:
        split_records = completion_data[completion_data['Split Invoice Description'].notna()]
        if not split_records.empty:
            print(f"\nSplit Invoice Summary:")
            print(f"  Total split invoice records: {len(split_records)}")
            
            # Group by original project to show split details
            if 'Original Amount' in completion_data.columns:
                # Count unique projects that were split
                unique_split_projects = split_records['ACGI #'].nunique()
                print(f"  Projects with split invoicing: {unique_split_projects}")
                
                # Show some examples
                print(f"  Example split invoices:")
                for acgi_num in split_records['ACGI #'].unique()[:3]:  # Show first 3 examples
                    project_splits = split_records[split_records['ACGI #'] == acgi_num]
                    if not project_splits.empty:
                        original_amount = project_splits.iloc[0].get('Original Amount', 0)
                        print(f"    {acgi_num}: ${original_amount:,.2f} split into {len(project_splits)} invoices")
                        for _, split_row in project_splits.iterrows():
                            desc = split_row.get('Split Invoice Description', 'N/A')
                            amount = split_row.get('Amount Invoiced', 0)
                            print(f"      - {desc}: ${amount:,.2f}")

def main():
    """
    Main function to update the quarterly YTD file.
    """
    # Get quarter information from user
    quarter_info = get_quarter_info()
    
    quarterly_file = quarter_info['quarterly_file']
    
    print("\n" + "="*60)
    print("QUARTERLY YTD UPDATER")
    print("="*60)
    print(f"Target file: {quarterly_file}")
    
    # Check if quarterly file exists
    if not os.path.exists(quarterly_file):
        print(f"Warning: {quarterly_file} not found!")
        print("A new file will be created.")
    
    # Collect completion data from all project lists
    completion_data = collect_completion_data(quarter_info)
    
    if completion_data.empty:
        print("No new completion data found. Nothing to update.")
        return
    
    # Print summary
    print_summary(completion_data)
    
    # Ask for confirmation
    response = input(f"\nDo you want to update {quarterly_file} with this data? (y/n): ").lower().strip()
    
    if response == 'y':
        success = update_quarterly_ytd(completion_data, quarter_info)
        if success:
            print(f"\n✓ {quarter_info['quarter_name']} {quarter_info['year']} YTD file updated successfully!")
        else:
            print(f"\n✗ Failed to update {quarter_info['quarter_name']} {quarter_info['year']} YTD file.")
    else:
        print("Update cancelled.")

if __name__ == "__main__":
    main() 