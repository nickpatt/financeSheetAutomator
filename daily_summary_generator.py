#!/usr/bin/env python3
"""
Daily Summary Generator with Quarterly YTD Updates
Generates daily invoicing summaries and updates quarterly YTD tracking.
"""

import pandas as pd
from datetime import datetime, timedelta
import os
import argparse
import sys
from openpyxl import load_workbook
from openpyxl.styles.colors import Color
from openpyxl.cell.cell import MergedCell

def get_quarter_from_date(date):
    """Determine which quarter a date falls into"""
    month = date.month
    if month in [1, 2, 3]:
        return 1
    elif month in [4, 5, 6]:
        return 2
    elif month in [7, 8, 9]:
        return 3
    else:  # month in [10, 11, 12]
        return 4

def get_quarter_name(quarter_num):
    """Get the ordinal name for a quarter number"""
    quarter_names = {1: "1st", 2: "2nd", 3: "3rd", 4: "4th"}
    return quarter_names.get(quarter_num, str(quarter_num))

def find_ytd_sheet(year, quarter_num):
    """Find the YTD sheet for the given year and quarter with new priority flow"""
    quarter_name = get_quarter_name(quarter_num)
    ytd_filename = f"{year} {quarter_name} Quarter YTD.xlsx"
    
    # Step 1: Check reports folder first
    reports_path = os.path.join('reports', ytd_filename)
    if os.path.exists(reports_path):
        print(f"✓ Found YTD sheet in reports folder: {reports_path}")
        return reports_path
    
    # Step 2: Not in reports, check N: drive and copy to quarterly sheets if found
    n_drive_path = os.path.join(rf'N:\Project List\{year} Project List', ytd_filename)
    if os.path.exists(n_drive_path):
        print(f"✓ Found YTD sheet on N: drive: {n_drive_path}")
        
        # Create quarterly sheets directory if it doesn't exist
        quarterly_sheets_dir = 'quarterly sheets'
        os.makedirs(quarterly_sheets_dir, exist_ok=True)
        
        # Copy from N: drive to quarterly sheets
        quarterly_sheets_path = os.path.join(quarterly_sheets_dir, ytd_filename)
        try:
            import shutil
            shutil.copy2(n_drive_path, quarterly_sheets_path)
            print(f"✓ Copied YTD sheet from N: drive to quarterly sheets: {quarterly_sheets_path}")
            return quarterly_sheets_path
        except Exception as e:
            print(f"✗ Failed to copy YTD sheet from N: drive: {e}")
            return None
    
    # Step 3: Check quarterly sheets as final fallback (in case it was already there)
    quarterly_sheets_path = os.path.join('quarterly sheets', ytd_filename)
    if os.path.exists(quarterly_sheets_path):
        print(f"✓ Found YTD sheet in quarterly sheets folder: {quarterly_sheets_path}")
        return quarterly_sheets_path
    
    print(f"⚠ YTD sheet not found in any location: {ytd_filename}")
    return None

def update_ytd_sheet_with_daily_table(target_date, daily_invoices_df):
    """Update the quarterly YTD sheet with the daily invoice table"""
    if daily_invoices_df.empty:
        print("No daily invoices to add to YTD sheet")
        return True
        
    year = target_date.year
    quarter_num = get_quarter_from_date(target_date)
    
    print(f"\n[INFO] Attempting to update YTD sheet for {year} Q{quarter_num}...")
    
    # Find the YTD sheet
    ytd_file_path = find_ytd_sheet(year, quarter_num)
    if not ytd_file_path:
        print(f"[ERROR] YTD sheet not found for {year} Q{quarter_num}. Aborting update.")
        return False
    else:
        print(f"[INFO] YTD sheet found: {ytd_file_path}")
    
    try:
        print(f"[DEBUG] Loading workbook: {ytd_file_path}")
        wb = load_workbook(ytd_file_path)
        ws = wb.active
        print(f"[DEBUG] Worksheet loaded. Max row: {ws.max_row}, Max column: {ws.max_column}")
        
        # Find where the date tables start (after the monthly summary rows and empty row)
        date_section_start = 5  # Start looking from row 5
        target_date_str = target_date.strftime('%A %m-%d-%Y')
        print(f"[DEBUG] Looking for date string: {target_date_str}")
        
        # Check if this date already exists
        existing_date_row = None
        current_row = date_section_start
        while current_row <= ws.max_row:
            cell_value = ws.cell(row=current_row, column=1).value
            print(f"[DEBUG] Row {current_row} Col 1 value: {cell_value}")
            if cell_value and target_date_str in str(cell_value):
                existing_date_row = current_row
                print(f"[INFO] Found existing date at row {current_row}")
                break
            current_row += 1
        
        if existing_date_row:
            # Replace existing table
            table_end_row = existing_date_row + 1
            while table_end_row <= ws.max_row:
                next_cell = ws.cell(row=table_end_row, column=1).value
                print(f"[DEBUG] Checking end of table at row {table_end_row}: {next_cell}")
                if next_cell and any(day in str(next_cell) for day in ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday', 'Saturday', 'Sunday']):
                    break
                if not next_cell and table_end_row > existing_date_row + 3:
                    break
                table_end_row += 1
            print(f"[INFO] Replacing table for {target_date_str} (rows {existing_date_row}-{table_end_row-1})")
            # Unmerge any merged cells in the range to be cleared
            print(f"[DEBUG] All merged ranges before clearing: {[str(rng) for rng in ws.merged_cells.ranges]}")
            
            # Direct approach: unmerge any merged range that contains the rows we want to clear
            for merged_range in list(ws.merged_cells.ranges):
                min_row, min_col, max_row, max_col = merged_range.bounds
                # If this merged range overlaps with any row we want to clear, unmerge it
                if min_row <= table_end_row and max_row >= existing_date_row:
                    print(f"[DEBUG] Unmerging merged range {str(merged_range)} (rows {min_row}-{max_row})")
                    try:
                        ws.unmerge_cells(str(merged_range))
                        print(f"[DEBUG] Successfully unmerged {str(merged_range)}")
                    except Exception as unmerge_err:
                        print(f"[ERROR] Failed to unmerge {str(merged_range)}: {unmerge_err}")
            
            # Print merged ranges after unmerging
            print(f"[DEBUG] All merged ranges after unmerging: {[str(rng) for rng in ws.merged_cells.ranges]}")
            
            # Force worksheet state update by reloading the worksheet object
            ws = wb.active
            for row in range(existing_date_row, table_end_row):
                for col in range(1, 11):
                    cell = ws.cell(row=row, column=col)
                    if isinstance(cell, MergedCell):
                        print(f"[ERROR] About to clear a MergedCell at row {row}, col {col} (should have been unmerged!)")
                        # Skip this cell and continue with the next one
                        continue
                    print(f"[DEBUG] Clearing cell at row {row}, col {col}")
                    try:
                        ws.cell(row=row, column=col).value = None
                    except Exception as clear_err:
                        print(f"[ERROR] Failed to clear cell at row {row}, col {col}: {clear_err}")
            insert_row = existing_date_row
        else:
            insert_row = ws.max_row + 1
            while insert_row > date_section_start and not any(ws.cell(row=insert_row-1, column=col).value for col in range(1, 11)):
                insert_row -= 1
            if insert_row > date_section_start:
                insert_row += 2
            print(f"[INFO] Adding new table for {target_date_str} at row {insert_row}")
        
        # Define styles for the table
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        day_style = Font(bold=True, color="FFFFFF")
        day_fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        total_style = Font(bold=True, color="FF0000")
        regular_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        center_align = Alignment(horizontal='center', vertical='center')
        current_row = insert_row
        # Add date header
        date_header = f"{target_date_str} (Invoice Date)"
        date_cell = ws.cell(row=current_row, column=1, value=date_header)
        date_cell.font = day_style
        date_cell.fill = day_fill
        date_cell.alignment = center_align
        date_cell.border = regular_border
        ws.merge_cells(f'A{current_row}:J{current_row}')
        print(f"[DEBUG] Wrote date header at row {current_row}")
        current_row += 1
        # Add column headers
        headers = ["ACGI Project / Invoice #", "Dept", "Project Number / Name", "Type", 
                  "Client / PO #", "Line #", "PO Date", "Amount", "Invoice Date", "Amount Invoiced"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = regular_border
            print(f"[DEBUG] Wrote header '{header}' at row {current_row}, col {col}")
        current_row += 1
        # Add data rows
        daily_total = 0
        for idx, (_, row) in enumerate(daily_invoices_df.iterrows()):
            values = [
                str(row.get('ACGI #', '')),
                str(row.get('Dept', '')),
                str(row.get('Project Number/Name', '')),
                str(row.get('Type', '')),
                str(row.get('Client / PO #', '')),
                str(row.get('Line # ', '')),
                row['PO Date'].strftime('%m/%d/%y') if pd.notna(row.get('PO Date')) else '',
                float(row.get('Amount', 0)) if pd.notna(row.get('Amount', 0)) else 0,
                row['Invoice Date'].strftime('%m/%d/%y') if pd.notna(row.get('Invoice Date')) else '',
                float(row.get('Amount Invoiced', 0)) if pd.notna(row.get('Amount Invoiced', 0)) else 0
            ]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = regular_border
                if col in [8, 10] and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
                    if col == 10:
                        daily_total += value
                print(f"[DEBUG] Wrote data at row {current_row}, col {col}: {value}")
            current_row += 1
        # Add total row
        total_cell = ws.cell(row=current_row, column=1, value="Total")
        total_cell.font = total_style
        total_cell.border = regular_border
        for col in range(2, 10):
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = regular_border
        amount_total_cell = ws.cell(row=current_row, column=10, value=daily_total)
        amount_total_cell.font = total_style
        amount_total_cell.number_format = '"$"#,##0.00'
        amount_total_cell.border = regular_border
        print(f"[DEBUG] Wrote total row at {current_row}, total: {daily_total}")
        # Save with conditional backup logic
        # Only create backup if file is NOT in reports folder
        is_in_reports = os.path.dirname(ytd_file_path).endswith('reports')
        
        try:
            if not is_in_reports:
                # Create backup for files in quarterly sheets folder
                backup_file = ytd_file_path.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
                wb.save(backup_file)
                print(f"[INFO] Backup saved: {backup_file}")
            else:
                print(f"[INFO] File in reports folder - no backup needed")
            
            wb.save(ytd_file_path)
            print(f"[SUCCESS] YTD sheet updated and saved: {ytd_file_path}")
        except Exception as save_err:
            print(f"[ERROR] Failed to save YTD sheet: {save_err}")
            return False
        action = "Updated existing" if existing_date_row else "Added new"
        print(f"[SUCCESS] {action} daily table in YTD sheet: {ytd_file_path}")
        print(f"  Date: {target_date_str}")
        print(f"  Records: {len(daily_invoices_df)}")
        return True
    except Exception as e:
        print(f"[ERROR] Exception while updating YTD sheet: {e}")
        return False

def get_user_input():
    """Get user input for date and other parameters"""
    print("=" * 60)
    print("Daily Summary Generator")
    print("=" * 60)
    
    # Get target date
    while True:
        date_input = input("\nEnter the target date (YYYY-MM-DD) or press Enter for today: ").strip()
        if not date_input:
            target_date = datetime.now().date()
            break
        try:
            target_date = datetime.strptime(date_input, "%Y-%m-%d").date()
            break
        except ValueError:
            print("Invalid date format. Please use YYYY-MM-DD (e.g., 2025-05-14)")
    
    # Note: No longer asking for base_dir as we use fixed N: drive with quarterly sheets fallback
    output_dir = input(f"\nEnter the output directory or press Enter for 'reports': ").strip()
    if not output_dir:
        output_dir = 'reports'
    
    print(f"\nConfiguration:")
    print(f"Target Date: {target_date}")
    print(f"Primary Data Directory: N:\\Project List\\")
    print(f"Fallback Data Directories: quarterly sheets, reports")
    print(f"Output Directory: {output_dir}")
    
    confirm = input("\nProceed with these settings? (y/n): ").strip().lower()
    if confirm != 'y':
        print("Cancelled.")
        sys.exit(0)
    
    return target_date, output_dir

def scan_available_project_files(start_year=2023, end_year=2030):
    """
    Scan for available Project List files from start_year to end_year.
    Returns a list of tuples (year, filepath) for files that exist.
    """
    available_files = []
    
    for year in range(start_year, end_year + 1):
        year_str = str(year)
        filename = f"{year_str} Project List.xlsx"
        file_path = find_file_in_locations(filename)
        
        if file_path:
            available_files.append((year_str, file_path))
            print(f"✓ Found {year_str} Project List")
        else:
            print(f"✗ {year_str} Project List not found")
    
    return available_files

def find_file_in_locations(filename):
    """
    Find a file in N:\Project List\, quarterly sheets folder, or reports folder.
    Tries both .xlsx and .xlsm extensions.
    Returns the full path if found, or None if not found in any location.
    """
    # Extract year from filename to build proper path structure
    year = None
    # Handle years 2023-2030
    for y in range(2023, 2031):
        if str(y) in filename:
            year = str(y)
            break
    
    # Try both .xlsx and .xlsm extensions
    base_filename = filename.replace('.xlsx', '').replace('.xlsm', '')
    extensions = ['.xlsx', '.xlsm']
    
    # Check N:\Project List\[year] Project List\
    if year:
        for ext in extensions:
            test_filename = base_filename + ext
            n_drive_path = os.path.join(rf'N:\Project List\{year} Project List', test_filename)
            if os.path.exists(n_drive_path):
                print(f"✓ Found {test_filename} in N:\\Project List\\{year} Project List\\")
                return n_drive_path
    
    # Check local quarterly sheets folder
    for ext in extensions:
        test_filename = base_filename + ext
        local_path = os.path.join('quarterly sheets', test_filename)
        if os.path.exists(local_path):
            print(f"✓ Found {test_filename} in quarterly sheets folder")
            return local_path
    
    # Check reports folder as backup
    for ext in extensions:
        test_filename = base_filename + ext
        reports_path = os.path.join('reports', test_filename)
        if os.path.exists(reports_path):
            print(f"✓ Found {test_filename} in reports folder")
            return reports_path
    
    # Not found in any location
    print(f"✗ {base_filename}.xlsx/.xlsm not found in N:\\Project List\\{year} Project List\\, quarterly sheets folder, or reports folder")
    return None

def collect_completion_data_for_quarter(base_dir, quarter_year=2025, quarter_num=2, selected_years=None):
    """
    Collect completion data for a specific quarter from all project lists.
    """
    # Use selected years if provided, otherwise default to 2023-2025
    if selected_years is None:
        selected_years = ['2023', '2024', '2025']
    
    project_files = []
    for year in selected_years:
        project_files.append((year, f'{year} Project List.xlsx'))
    
    project_lists = []
    for year, filename in project_files:
        file_path = find_file_in_locations(filename)
        if file_path:
            project_lists.append((year, file_path))
        else:
            print(f"Warning: Could not find {filename} for quarter data collection")
            continue
    
    # Define quarter date ranges
    quarter_ranges = {
        1: (datetime(quarter_year, 1, 1), datetime(quarter_year, 3, 31)),
        2: (datetime(quarter_year, 4, 1), datetime(quarter_year, 6, 30)),
        3: (datetime(quarter_year, 7, 1), datetime(quarter_year, 9, 30)),
        4: (datetime(quarter_year, 10, 1), datetime(quarter_year, 12, 31)),
    }
    
    if quarter_num not in quarter_ranges:
        print(f"Invalid quarter number: {quarter_num}")
        return pd.DataFrame()
    
    q_start, q_end = quarter_ranges[quarter_num]
    print(f"Collecting completion data for Q{quarter_num} {quarter_year} ({q_start.strftime('%Y-%m-%d')} to {q_end.strftime('%Y-%m-%d')})...")
    
    all_completion_data = []
    
    for year, file_path in project_lists:
        try:
            df = pd.read_excel(file_path, sheet_name=year, header=5)
            
            # Get the correct column names (handle variations)
            acgi_col = None
            for col in df.columns:
                if 'acgi' in str(col).lower() and '#' in str(col):
                    acgi_col = col
                    break
            
            if acgi_col is None:
                acgi_col = df.columns[0]  # Use first column as fallback
            
            completion_data = df[['Completion Date', 'Amount Invoiced', acgi_col, 'Project Number/Name', 'Client / PO #']].copy()
            completion_data = completion_data.rename(columns={acgi_col: 'ACGI #'})
            
            # Clean the data
            completion_data = completion_data.dropna(subset=['Completion Date', 'Amount Invoiced'])
            completion_data['Completion Date'] = pd.to_datetime(completion_data['Completion Date'])
            completion_data['Amount Invoiced'] = pd.to_numeric(completion_data['Amount Invoiced'], errors='coerce')
            
            # Filter for quarter dates
            quarter_data = completion_data[
                (completion_data['Completion Date'] >= q_start) & 
                (completion_data['Completion Date'] <= q_end)
            ].copy()
            
            if not quarter_data.empty:
                quarter_data['Source_Year'] = year
                all_completion_data.append(quarter_data)
                print(f"Found {len(quarter_data)} Q{quarter_num} {quarter_year} completion records from {year}")
            
        except Exception as e:
            print(f"Error processing {year} Project List: {e}")
    
    if all_completion_data:
        combined_data = pd.concat(all_completion_data, ignore_index=True)
        return combined_data
    else:
        return pd.DataFrame()

def update_quarterly_ytd_file(base_dir, completion_data, quarter_year=2025, quarter_num=2):
    """
    Update the quarterly YTD Excel file with completion data.
    """
    quarterly_file = os.path.join(base_dir, f'{quarter_year} {quarter_num}{"nd" if quarter_num == 2 else ("st" if quarter_num == 1 else ("rd" if quarter_num == 3 else "th"))} Quarter YTD.xlsx')
    
    if not os.path.exists(quarterly_file):
        print(f"Quarterly file not found: {quarterly_file}")
        return False
    
    try:
        wb = load_workbook(quarterly_file)
        ws = wb.active
        
        # Calculate monthly totals
        if not completion_data.empty:
            completion_data['Month'] = completion_data['Completion Date'].dt.month
            monthly_totals = completion_data.groupby('Month')['Amount Invoiced'].sum()
            
            # Map months to column numbers (assuming standard layout)
            month_columns = {1: 1, 2: 2, 3: 3, 4: 4, 5: 5, 6: 6, 7: 7, 8: 8, 9: 9, 10: 10, 11: 11, 12: 12}
            
            # Update monthly totals in row 2
            for month_num, total in monthly_totals.items():
                if month_num in month_columns:
                    col_num = month_columns[month_num]
                    current_value = ws.cell(row=2, column=col_num).value or 0
                    new_value = current_value + total
                    ws.cell(row=2, column=col_num, value=new_value)
            
            # Find last row and add new records
            last_row = ws.max_row
            while last_row > 1 and all(ws.cell(row=last_row, column=col).value is None for col in range(1, 11)):
                last_row -= 1
            
            # Add completion records
            for idx, row in completion_data.iterrows():
                new_row = last_row + 1 + idx
                ws.cell(row=new_row, column=1, value=row.get('ACGI #', ''))
                ws.cell(row=new_row, column=2, value='')  # Dept
                ws.cell(row=new_row, column=3, value=row.get('Project Number/Name', ''))
                ws.cell(row=new_row, column=4, value='Completion')
                ws.cell(row=new_row, column=5, value=row.get('Client / PO #', ''))
                ws.cell(row=new_row, column=6, value='')  # Line #
                ws.cell(row=new_row, column=7, value='')  # PO Date
                ws.cell(row=new_row, column=8, value=row['Amount Invoiced'])
                ws.cell(row=new_row, column=9, value=row['Completion Date'])
                ws.cell(row=new_row, column=10, value=row['Amount Invoiced'])
        
        # Save with backup
        backup_file = quarterly_file.replace('.xlsx', f'_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx')
        wb.save(backup_file)
        wb.save(quarterly_file)
        
        print(f"✓ Updated quarterly YTD file: {quarterly_file}")
        return True
        
    except Exception as e:
        print(f"Error updating quarterly YTD file: {e}")
        return False



def generate_summary(target_date, output_dir, selected_years=None):
    """Generate the daily summary report"""
    
    print(f"\nGenerating summary for {target_date}...")
    
    # --- Configuration with Fallback Logic ---
    print("Locating required files...")
    
    # Find project list files with fallback
    invoice_sources = []
    
    # Use selected years if provided, otherwise default to 2023-2025
    if selected_years is None:
        selected_years = ['2023', '2024', '2025']
    
    project_files = []
    for year in selected_years:
        project_files.append((year, f'{year} Project List.xlsx'))
    
    for year, filename in project_files:
        file_path = find_file_in_locations(filename)
        if file_path:
            invoice_sources.append((year, file_path))
        else:
            print(f"Error: Could not find {filename}")
            return False
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    print(f"✓ Output directory ready: {output_dir}")
    
    try:
        # --- 1) Load and combine all Amount Invoiced entries for date-based totals ---
        print("Loading invoice data...")
        inv_dfs = []
        for year, path in invoice_sources:
            df = pd.read_excel(path, sheet_name=year, header=5)
            df['Invoice Date'] = pd.to_datetime(df['Invoice Date'], errors='coerce').dt.date
            inv_dfs.append(df)
        invoices = pd.concat(inv_dfs, ignore_index=True)
        
        # --- 2) Compute periods ---
        week_start  = target_date - timedelta(days=target_date.weekday())
        month_start = target_date.replace(day=1)
        
        today_total = invoices.loc[invoices['Invoice Date']==target_date, 'Amount'].sum()
        invoice_total = invoices.loc[invoices['Invoice Date']==target_date, 'Amount Invoiced'].sum()
        week_total  = invoices.loc[(invoices['Invoice Date']>=week_start) & (invoices['Invoice Date']<=target_date), 'Amount Invoiced'].sum()
        month_total = invoices.loc[(invoices['Invoice Date']>=month_start) & (invoices['Invoice Date']<=target_date), 'Amount Invoiced'].sum()
        
        # --- 3) Load vendor payments from the Project List for the target year ---
        print("Loading vendor payment data from Project List...")

        target_year = str(target_date.year)
        project_file_dict = {year: path for year, path in invoice_sources}

        if target_year not in project_file_dict:
            print(f"Error: Project List file for year {target_year} not found.")
            return False

        # Read the Project List file for the target year
        try:
            df = pd.read_excel(project_file_dict[target_year], sheet_name=target_year, header=None)
            # Find the last two non-empty rows in column G
            non_empty_rows = []
            for idx in range(len(df)):
                if pd.notna(df.iloc[idx, 6]) and str(df.iloc[idx, 6]).strip():
                    non_empty_rows.append(idx)
            if len(non_empty_rows) < 2:
                print(f"Error: Could not find enough non-empty rows in {target_year} Project List.")
                return False
            to_invoice_row = non_empty_rows[-2]
            # Get vendor payment value from column M (index 12)
            target_year_vendor_payment = float(df.iloc[to_invoice_row, 12])
            print(f"Vendor payments for {target_year} (to invoice row): ${target_year_vendor_payment:,.2f}")
        except Exception as e:
            print(f"Error reading vendor payments from {target_year} Project List: {e}")
            return False
        
        # --- 4) Get receivables data from Project List files ---
        print("Processing project list files for receivables data...")
        years = selected_years
        recv_by_year = []
        pay_by_year = []
        year_data = {}
        
        # Use the already found project list files
        project_file_dict = {year: path for year, path in invoice_sources}
        
        for year in years:
            if year not in project_file_dict:
                print(f"Warning: {year} project list not available for receivables processing")
                recv_by_year.append(0)
                pay_by_year.append(0)
                continue
                
            file_path = project_file_dict[year]
            df = pd.read_excel(file_path, sheet_name=year, header=None)
            
            # Find the last two non-empty rows in column G
            non_empty_rows = []
            for idx in range(len(df)):
                if pd.notna(df.iloc[idx, 6]) and str(df.iloc[idx, 6]).strip():
                    non_empty_rows.append(idx)
            
            if len(non_empty_rows) >= 2:
                to_invoice_row = non_empty_rows[-2]
                less_hold_row = non_empty_rows[-1]
                totals_row = to_invoice_row - 1
                
                print(f"Found rows in {year}:")
                print(f"  Totals row {totals_row}: {str(df.iloc[totals_row, 6]).strip()}")
                print(f"  To Invoice row {to_invoice_row}: {str(df.iloc[to_invoice_row, 6]).strip()}")
                print(f"  Less hold row {less_hold_row}: {str(df.iloc[less_hold_row, 6]).strip()}")
                
                year_data[year] = {
                    'totals_row': totals_row,
                    'to_invoice_row': to_invoice_row,
                    'less_hold_row': less_hold_row
                }
                
                try:
                    recv_amount = float(df.iloc[totals_row, 12])  # Column M - one row higher than to_invoice_row
                    
                    # Sum all amounts in vendor payment column for vendors to be paid - ONLY light blue/aqua colored cells
                    # 2023 & 2024 use Column V (22), 2025 uses Column W (23)
                    pay_amount = 0
                    colored_cells_count = 0
                    
                    # Use openpyxl to check cell colors
                    
                    wb = load_workbook(file_path)
                    ws = wb[year] if year in wb.sheetnames else wb.active
                    
                    # Determine which column to use based on year
                    if year in ['2023', '2024']:
                        vendor_column = 22  # Column V (1-indexed for openpyxl)
                        column_name = 'V'
                    else:  # 2025 and later
                        vendor_column = 23  # Column W (1-indexed for openpyxl)
                        column_name = 'W'
                    
                    print(f"  {year} - Using Column {column_name} for vendor payments")
                    
                    for row_num in range(1, ws.max_row + 1):
                        cell = ws.cell(row=row_num, column=vendor_column)
                        cell_value = cell.value
                        
                        if cell_value is not None and str(cell_value).strip():
                            # Check if cell has a fill color
                            if cell.fill and cell.fill.start_color and cell.fill.start_color.rgb:
                                color_rgb = cell.fill.start_color.rgb
                                
                                # Define the target cyan/aqua color: rgb(3,255,255) = 03FFFF
                                # and similar cyan/aqua variations
                                target_cyan_colors = [
                                    '03FFFF',  # rgb(3,255,255) - Target color
                                    '00FFFF',  # rgb(0,255,255) - Pure Cyan
                                    '01FFFF',  # rgb(1,255,255) 
                                    '02FFFF',  # rgb(2,255,255)
                                    '04FFFF',  # rgb(4,255,255)
                                    '05FFFF',  # rgb(5,255,255)
                                    '06FFFF',  # rgb(6,255,255)
                                    '07FFFF',  # rgb(7,255,255)
                                    '08FFFF',  # rgb(8,255,255)
                                    '09FFFF',  # rgb(9,255,255)
                                    '0AFFFF',  # rgb(10,255,255)
                                ]
                                
                                # Check if the cell color matches the target cyan color or close variants
                                color_match = False
                                if color_rgb.upper() in [color.upper() for color in target_cyan_colors]:
                                    color_match = True
                                elif color_rgb.upper().startswith('FF') and len(color_rgb) == 8:
                                    # Handle ARGB format (FF + RGB)
                                    rgb_part = color_rgb[2:]
                                    if rgb_part.upper() in [color.upper() for color in target_cyan_colors]:
                                        color_match = True
                                
                                if color_match:
                                    try:
                                        numeric_value = float(cell_value)
                                        pay_amount += numeric_value
                                        colored_cells_count += 1
                                        print(f"    Added {year} Row {row_num} (color {color_rgb}): ${numeric_value:,.2f}")
                                    except (ValueError, TypeError):
                                        continue  # Skip non-numeric values
                    
                    wb.close()
                    print(f"  {year} - Column {column_name} cyan cells total: ${pay_amount:,.2f} ({colored_cells_count} cells)")
                    
                    recv_by_year.append(recv_amount)
                    pay_by_year.append(pay_amount)
                except Exception as e:
                    print(f"Warning: Error reading data for {year}: {e}")
                    recv_by_year.append(0)
                    pay_by_year.append(0)
            else:
                print(f"Warning: Could not find enough non-empty rows in {year} Project List")
                recv_by_year.append(0)
                pay_by_year.append(0)
        
        # Calculate totals
        total_rec = sum(recv_by_year)
        total_pay = sum(pay_by_year)
        net_receivables = total_rec - total_pay
        
        # --- 5) Create Excel file with all tables in one sheet ---
        print("Creating Excel file with tables...")
        excel_file = os.path.join(output_dir, f'daily_summary_tables_{target_date.strftime("%Y%m%d")}.xlsx')
        
        # Import openpyxl components
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
        from openpyxl.utils import get_column_letter
        
        # Create new workbook
        wb = Workbook()
        ws = wb.active
        ws.title = 'Daily Summary Tables'
        
        # Define styles (same as quarterly_ytd_updater)
        # Header style (bold, gray background)
        header_style = Font(bold=True)
        header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        
        # Title style (bold, green background)
        title_style = Font(bold=True, color="FFFFFF")
        title_fill = PatternFill(start_color="00AA00", end_color="00AA00", fill_type="solid")
        
        # Total style (bold, red font)
        total_style = Font(bold=True, color="FF0000")
        
        # Regular border
        regular_border = Border(
            left=Side(border_style="thin", color="000000"),
            right=Side(border_style="thin", color="000000"),
            top=Side(border_style="thin", color="000000"),
            bottom=Side(border_style="thin", color="000000")
        )
        
        # Center alignment
        center_align = Alignment(horizontal='center', vertical='center')
        
        current_row = 1
        
        # --- ADD SUMMARY SECTION AT TOP ---
        # Main title
        summary_title = f"Daily Invoicing Summary - {target_date.strftime('%A, %B %d, %Y')}"
        title_cell = ws.cell(row=current_row, column=1, value=summary_title)
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
        title_cell.alignment = center_align
        title_cell.border = regular_border
        ws.merge_cells(f'A{current_row}:J{current_row}')
        current_row += 2
        
        # Summary data rows
        summary_data = [
            (f"Today ({target_date})", f"${today_total:,.2f}"),
            (f"This Week (since {week_start})", f"${week_total:,.2f}"),
            (f"This Month (since {month_start})", f"${month_total:,.2f}"),
            ("Total Payments Received", f"${invoice_total:,.2f}"),
            ("Total Receivables", f"${total_rec:,.2f}"),
            ("Vendors to be paid", f"${total_pay:,.2f}"),
            ("Net Receivables", f"${net_receivables:,.2f}")
        ]
        
        for label, value in summary_data:
            # Label in column A
            label_cell = ws.cell(row=current_row, column=1, value=label)
            label_cell.font = Font(bold=True)
            label_cell.border = regular_border
            
            # Value in column B
            value_cell = ws.cell(row=current_row, column=2, value=value)
            value_cell.font = Font(bold=True, color="0000AA")
            value_cell.border = regular_border
            
            # Empty cells for formatting consistency
            for col in range(3, 11):
                ws.cell(row=current_row, column=col, value="").border = regular_border
            
            current_row += 1
        
        # Add spacing after summary
        current_row += 3
        
        # Table 1: Invoice Details
        daily_inv = invoices[invoices['Invoice Date']==target_date]
        
        # --- Update YTD Sheet with Daily Table ---
        print("\n" + "="*50)
        print("UPDATING QUARTERLY YTD SHEET")
        print("="*50)
        
        ytd_success = update_ytd_sheet_with_daily_table(target_date, daily_inv)
        if ytd_success:
            print("✓ YTD sheet updated successfully")
        else:
            print("⚠ YTD sheet update failed or skipped")
        
        print("="*50)
        print("CONTINUING WITH DAILY SUMMARY GENERATION")
        print("="*50 + "\n")
        
        # Title row
        ws.cell(row=current_row, column=1, value=f"Invoices for {target_date.strftime('%A %m-%d-%Y')}")
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = title_style
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        title_cell.border = regular_border
        # Merge title across columns
        ws.merge_cells(f'A{current_row}:J{current_row}')
        current_row += 1
        
        # Column headers
        headers = ["ACGI Project / Invoice #", "Dept", "Project Number / Name", "Type", 
                  "Client / PO #", "Line #", "PO Date", "Amount", "Invoice Date", "Amount Invoiced"]
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = regular_border
        current_row += 1
        
        # Data rows
        daily_total = 0
        
        for _, row in daily_inv.iterrows():
            # Column mapping
            values = [
                str(row.get('ACGI #', '')),
                str(row.get('Dept', '')),
                str(row.get('Project Number/Name', '')),
                str(row.get('Type', '')),
                str(row.get('Client / PO #', '')),
                str(row.get('Line # ', '')),
                row['PO Date'].strftime('%m/%d/%y') if pd.notna(row.get('PO Date')) else '',
                float(row.get('Amount', 0)) if pd.notna(row.get('Amount')) else 0,
                row['Invoice Date'].strftime('%m/%d/%y') if pd.notna(row.get('Invoice Date')) else '',
                float(row.get('Amount Invoiced', 0)) if pd.notna(row.get('Amount Invoiced')) else 0
            ]
            
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=current_row, column=col, value=value)
                cell.border = regular_border
                # Format currency columns
                if col in [8, 10] and isinstance(value, (int, float)):
                    cell.number_format = '"$"#,##0.00'
                    daily_total += value if col == 10 else 0
            
            current_row += 1
        
        # Total row
        total_cell = ws.cell(row=current_row, column=1, value="Total")
        total_cell.font = total_style
        total_cell.border = regular_border
        
        for col in range(2, 10):
            cell = ws.cell(row=current_row, column=col, value="")
            cell.border = regular_border
        
        amount_total_cell = ws.cell(row=current_row, column=10, value=daily_total)
        amount_total_cell.font = total_style
        amount_total_cell.number_format = '"$"#,##0.00'
        amount_total_cell.border = regular_border
        
        current_row += 1
        print("✓ Added Invoice Details table")
        
        # Add spacing between tables
        current_row += 3
        
        # Table 2: Receivables vs Vendors
        # Title row
        ws.cell(row=current_row, column=1, value="Receivables vs. Vendors to be Paid by Year")
        title_cell = ws.cell(row=current_row, column=1)
        title_cell.font = title_style
        title_cell.fill = title_fill
        title_cell.alignment = center_align
        title_cell.border = regular_border
        ws.merge_cells(f'A{current_row}:C{current_row}')
        current_row += 1
        
        # Headers
        headers = ['Year', 'Receivables', 'Vendors to be paid']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_style
            cell.fill = header_fill
            cell.alignment = center_align
            cell.border = regular_border
        current_row += 1
        
        # Data rows
        for yr, rcv, pay in zip(years, recv_by_year, pay_by_year):
            ws.cell(row=current_row, column=1, value=str(yr)).border = regular_border
            
            rcv_cell = ws.cell(row=current_row, column=2, value=rcv)
            rcv_cell.number_format = '"$"#,##0.00'
            rcv_cell.border = regular_border
            
            pay_cell = ws.cell(row=current_row, column=3, value=pay)
            pay_cell.number_format = '"$"#,##0.00'
            pay_cell.border = regular_border
            
            current_row += 1
        
        # Total row
        total_cell = ws.cell(row=current_row, column=1, value="Total")
        total_cell.font = total_style
        total_cell.border = regular_border
        
        total_rcv_cell = ws.cell(row=current_row, column=2, value=total_rec)
        total_rcv_cell.font = total_style
        total_rcv_cell.number_format = '"$"#,##0.00'
        total_rcv_cell.border = regular_border
        
        total_pay_cell = ws.cell(row=current_row, column=3, value=total_pay)
        total_pay_cell.font = total_style
        total_pay_cell.number_format = '"$"#,##0.00'
        total_pay_cell.border = regular_border
        
        current_row += 1
        print("✓ Added Receivables vs Vendors table")
        
        # Table 3-5: Year Details
        for year in ['2023', '2024', '2025']:
            if year not in year_data or year not in project_file_dict:
                print(f"⚠ Skipping {year} Details - data not available")
                continue
                
            # Add spacing between tables
            current_row += 3
            
            # Title row
            ws.cell(row=current_row, column=1, value=f'{year} Details')
            title_cell = ws.cell(row=current_row, column=1)
            title_cell.font = title_style
            title_cell.fill = title_fill
            title_cell.alignment = center_align
            title_cell.border = regular_border
            ws.merge_cells(f'A{current_row}:G{current_row}')
            current_row += 1
            
            file_path = project_file_dict[year]
            try:
                df = pd.read_excel(file_path, sheet_name=year, header=None)
                rows = year_data[year]
                
                # Row 1 (totals row)
                try:
                    ws.cell(row=current_row, column=2, value=float(df.iloc[rows['totals_row'], 7]))
                    ws.cell(row=current_row, column=4, value=float(df.iloc[rows['totals_row'], 9]))
                    ws.cell(row=current_row, column=5, value=float(df.iloc[rows['totals_row'], 10]))
                    ws.cell(row=current_row, column=7, value=float(df.iloc[rows['totals_row'], 12]))
                    
                    # Format currency
                    for col in [2, 4, 5, 7]:
                        cell = ws.cell(row=current_row, column=col)
                        cell.number_format = '"$"#,##0.00'
                        cell.border = regular_border
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Warning: Could not convert some values in {year} totals row: {e}")
                
                # Add borders to empty cells in row 1
                for col in range(1, 8):
                    if ws.cell(row=current_row, column=col).value is None:
                        ws.cell(row=current_row, column=col, value="").border = regular_border
                current_row += 1
                
                # Row 2 (to invoice row)
                ws.cell(row=current_row, column=1, value="To Invoice").border = regular_border
                try:
                    cell2 = ws.cell(row=current_row, column=2, value=float(df.iloc[rows['to_invoice_row'], 7]))
                    cell2.number_format = '"$"#,##0.00'
                    cell2.border = regular_border
                    
                    cell7 = ws.cell(row=current_row, column=7, value=float(df.iloc[rows['to_invoice_row'], 12]))
                    cell7.number_format = '"$"#,##0.00'
                    cell7.border = regular_border
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Warning: Could not convert some values in {year} to invoice row: {e}")
                
                # Add borders to empty cells in row 2
                for col in range(1, 8):
                    if ws.cell(row=current_row, column=col).value is None:
                        ws.cell(row=current_row, column=col, value="").border = regular_border
                current_row += 1
                
                # Row 3 (less hold row)
                ws.cell(row=current_row, column=1, value="To invoice less hold").border = regular_border
                try:
                    cell2 = ws.cell(row=current_row, column=2, value=float(df.iloc[rows['less_hold_row'], 7]))
                    cell2.number_format = '"$"#,##0.00'
                    cell2.border = regular_border
                except (ValueError, TypeError, IndexError) as e:
                    print(f"Warning: Could not convert some values in {year} less hold row: {e}")
                
                # Add borders to empty cells in row 3
                for col in range(1, 8):
                    if ws.cell(row=current_row, column=col).value is None:
                        ws.cell(row=current_row, column=col, value="").border = regular_border
                current_row += 1
                
                print(f"✓ Added {year} Details table")
                
            except Exception as e:
                print(f"Error processing {year} Details: {e}")
                current_row += 1
        
        # Auto-adjust column widths for the entire sheet
        for col in range(1, 11):  # Covers all columns used
            column_letter = get_column_letter(col)
            max_length = 0
            for row in ws[column_letter]:
                try:
                    if len(str(row.value)) > max_length:
                        max_length = len(str(row.value))
                except:
                    pass
            adjusted_width = max(max_length + 2, 12)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save the workbook
        try:
            wb.save(excel_file)
            print(f"✓ Successfully created Excel file: {excel_file}")
        except Exception as e:
            print(f"Error saving Excel file: {e}")
            return False
        
        print(f"\n✓ Summary generated successfully!")
        print(f"  Excel File (with summary and tables): {excel_file}")
        if ytd_success:
            quarter_num = get_quarter_from_date(target_date)
            print(f"  YTD Sheet: Updated {target_date.year} Q{quarter_num} Quarter YTD")
        print(f"  Date: {target_date}")
        print(f"  Today's Total: ${today_total:,.2f}")
        print(f"  Total Payments Received: ${daily_total:,.2f}")
        print(f"  Week Total: ${week_total:,.2f}")
        print(f"  Month Total: ${month_total:,.2f}")
        
        return True
        
    except Exception as e:
        print(f"Error generating summary: {str(e)}")
        return False

def main():
    """Main function"""
    parser = argparse.ArgumentParser(description='Generate daily invoicing summary reports')
    parser.add_argument('--date', '-d', help='Target date (YYYY-MM-DD)')
    parser.add_argument('--data-dir', help='Data directory path (default: fwytdreport)')
    parser.add_argument('--output-dir', help='Output directory path (default: reports)')
    parser.add_argument('--years', nargs='+', help='Years to process (e.g., --years 2023 2024 2025)')
    parser.add_argument('--scan-files', action='store_true', help='Scan for available Project List files')
    parser.add_argument('--interactive', '-i', action='store_true', help='Interactive mode')
    parser.add_argument('--update-ytd', action='store_true', help='Update quarterly YTD file')
    parser.add_argument('--quarter', type=int, default=2, help='Quarter number (1-4)')
    parser.add_argument('--year', type=int, default=2025, help='Year for quarterly update')
    
    args = parser.parse_args()
    
    # Handle scan files option
    if args.scan_files:
        print("Scanning for available Project List files...")
        available_files = scan_available_project_files(2023, 2030)
        if available_files:
            print(f"Found {len(available_files)} available files:")
            for year, path in available_files:
                print(f"  {year}: {path}")
        else:
            print("No Project List files found in any location.")
        return
    
    if args.interactive or not any([args.date, args.data_dir, args.output_dir]):
        # Interactive mode
        target_date, output_dir = get_user_input()
        selected_years = None  # Use default years in interactive mode
    else:
        # Command line mode
        if args.date:
            try:
                target_date = datetime.strptime(args.date, "%Y-%m-%d").date()
            except ValueError:
                print("Error: Invalid date format. Use YYYY-MM-DD")
                sys.exit(1)
        else:
            target_date = datetime.now().date()
        
        output_dir = args.output_dir or 'reports'
        
        # Validate and convert years to strings
        if args.years:
            selected_years = []
            for year in args.years:
                try:
                    year_int = int(year)
                    if 2023 <= year_int <= 2030:
                        selected_years.append(str(year_int))
                    else:
                        print(f"Error: Year {year} is not in the supported range (2023-2030)")
                        sys.exit(1)
                except ValueError:
                    print(f"Error: '{year}' is not a valid year")
                    sys.exit(1)
        else:
            selected_years = None  # Use defaults
    
    # Generate the summary
    success = generate_summary(target_date, output_dir, selected_years)
    
    if success:
        print("\nReport generation completed successfully!")
    else:
        print("\nReport generation failed!")
        sys.exit(1)

    # Update quarterly YTD if requested
    if args.update_ytd:
        print(f"\nUpdating quarterly YTD for Q{args.quarter} {args.year}...")
        completion_data = collect_completion_data_for_quarter(output_dir, args.year, args.quarter)
        
        if not completion_data.empty:
            print(f"Found {len(completion_data)} completion records for Q{args.quarter} {args.year}")
            success = update_quarterly_ytd_file(output_dir, completion_data, args.year, args.quarter)
            if success:
                print("✓ Quarterly YTD update completed")
            else:
                print("✗ Quarterly YTD update failed")
        else:
            print("No completion data found for the specified quarter")

if __name__ == "__main__":
    main() 